"""
Data consolidation into master report
"""

import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..config.cell_mappings import FIELD_DISPLAY_NAMES, CELL_MAPPINGS

logger = logging.getLogger(__name__)


class DataConsolidator:
    """Consolidate extracted data into master report"""
    

    def create_report(self, results, output_folder):
        """Create consolidated Excel report with multiple custom sheets"""
        try:
            # Import the new configuration
            from ..config.cell_mappings import EXTRACTION_CONFIG, EXTRACTION_OPTIONS
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            section_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Process each configured sheet
            for sheet_name, sheet_config in EXTRACTION_CONFIG.items():
                logger.info(f"Creating sheet: {sheet_name}")
                
                # Create new sheet
                ws = wb.create_sheet(sheet_name)
                
                # Add description if available
                start_row = 1
                if 'description' in sheet_config:
                    descriptions = sheet_config['description']
                    for i, desc_line in enumerate(descriptions):
                        ws[f'A{i+1}'] = desc_line
                        if i == 0:  # First line - make it bold and larger
                            ws[f'A{i+1}'].font = Font(bold=True, size=14)
                        else:  # Other lines - normal but italic
                            ws[f'A{i+1}'].font = Font(italic=True, size=11)
                        # Merge across multiple columns for readability
                        ws.merge_cells(f'A{i+1}:J{i+1}')
                    
                    # Add blank row after description
                    start_row = len(descriptions) + 2
                else:
                    # No description - use original title
                    ws['A1'] = f"{sheet_name} Report"
                    ws['A1'].font = section_font
                    ws.merge_cells('A1:C1')
                    start_row = 3

                # Build headers
                headers = ['Client Name', 'Successfully Run']
                
                # Add configured column headers
                for mapping in sheet_config['mappings']:
                    headers.append(mapping['output_column'])
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Write data for each client
                current_row = start_row + 1
                report_type = sheet_config['report_type']
                
                for result in results:
                    # Client name
                    ws.cell(row=current_row, column=1, value=result['client']).border = border
                    
                    # Success status
                    report_data = result.get(report_type.lower(), {})
                    is_success = report_data.get('status', {}).get('success', False)
                    
                    status_cell = ws.cell(
                        row=current_row, 
                        column=2, 
                        value='✓' if is_success else '✗'
                    )
                    status_cell.border = border
                    status_cell.font = Font(
                        color="107C10" if is_success else "D83B01", 
                        bold=True
                    )
                    
                    # Process each mapping
                    col_index = 3
                    for mapping in sheet_config['mappings']:
                        value = None
                        
                        if not is_success:
                            value = EXTRACTION_OPTIONS.get('missing_data_text', 'Missing')
                        elif mapping.get('calculation'):
                            # Handle calculated fields
                            value = self._calculate_value(
                                report_data.get('extracted_values', {}),
                                mapping['calculation'],
                                wb  # Pass workbook for access to other data
                            )
                        else:
                            # Get direct value
                            extracted_data = report_data.get('data', {})
                            
                            # Find matching value in extracted data
                            for field_name, field_value in extracted_data.items():
                                if mapping['output_column'] in field_name:
                                    value = field_value
                                    break
                            
                            # Fallback: try exact match
                            if value is None and mapping['input_cell']:
                                key = f"{sheet_name}_{mapping['output_column']}".replace(' ', '_')
                                value = extracted_data.get(key)
                        
                        # Handle missing values
                        if value is None or value == '':
                            value = EXTRACTION_OPTIONS.get('missing_data_text', 'Missing')
                        
                        # Write value
                        cell = ws.cell(row=current_row, column=col_index, value=value)
                        cell.border = border
                        
                        # Format numbers
                        if isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'
                        
                        col_index += 1
                    
                    current_row += 1
                
                # Auto-fit columns for this sheet
                self._autofit_columns(ws)
            
            # Add summary sheet
            self._create_summary_sheet(wb, results)
            
            # Save report
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            report_name = f"PQ_Extraction_Report_{timestamp}.xlsx"
            report_path = output_folder / report_name
            
            wb.save(report_path)
            logger.info(f"Consolidated report saved: {report_path}")
            
            return report_path
            
        except Exception as e:
            logger.error(f"Error creating report: {e}", exc_info=True)
            raise
    
    def _calculate_value(self, extracted_values, calculation, workbook):
        """Calculate value based on configuration"""
        try:
            calc_type = calculation.get('type')
            
            # Get values
            value1 = extracted_values.get(calculation.get('cell1'), 0)
            value2 = extracted_values.get(calculation.get('cell2'), 0)
            
            # Convert to numbers
            try:
                num1 = float(value1) if value1 else 0
                num2 = float(value2) if value2 else 0
            except (TypeError, ValueError):
                return 'Error'
            
            # Perform calculation
            if calc_type == 'subtract':
                return num1 - num2
            elif calc_type == 'add':
                return num1 + num2
            elif calc_type == 'multiply':
                return num1 * num2
            elif calc_type == 'divide':
                return num1 / num2 if num2 != 0 else 'Div/0'
            else:
                return 'Unknown'
                
        except Exception as e:
            logger.error(f"Calculation error: {e}")
            return 'Error'
    
    def _create_summary_sheet(self, wb, results):
        """Create summary sheet with extraction statistics"""
        from ..config.cell_mappings import EXTRACTION_CONFIG

        ws = wb.create_sheet("Summary", 0)  # Insert at beginning
        
        # Title
        ws['A1'] = "Extraction Summary"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Clients: {len(results)}"
        
        # Count successes
        itc_success = sum(1 for r in results if r.get('itc', {}).get('status', {}).get('success'))
        sales_success = sum(1 for r in results if r.get('sales', {}).get('status', {}).get('success'))
        
        ws['A5'] = f"ITC Reports Processed: {itc_success}/{len(results)}"
        ws['A6'] = f"Sales Reports Processed: {sales_success}/{len(results)}"
        
        # List of sheets created
        ws['A8'] = "Sheets Created:"
        ws['A8'].font = Font(bold=True)
        
        row = 9
        for sheet_name in EXTRACTION_CONFIG.keys():
            ws[f'A{row}'] = f"• {sheet_name}"
            row += 1
    
    def _write_section(self, ws, results, report_key, report_type, start_row, 
                      header_font, header_fill, border, subsection_font):
        """Write a section (ITC or Sales) to worksheet"""
        
        # Get all possible fields from CELL_MAPPINGS
        all_fields = set()
        mappings = CELL_MAPPINGS.get(report_type, {})
        for sheet_cells in mappings.values():
            all_fields.update(sheet_cells.values())
        
        # Also get fields from actual results
        for result in results:
            data = result[report_key].get('data', {})
            all_fields.update(data.keys())
        
        all_fields = sorted(list(all_fields))
        
        if not all_fields:
            # No fields defined, show message
            ws[f'A{start_row}'] = f"No cell mappings defined for {report_type} reports"
            ws[f'A{start_row}'].font = Font(italic=True, color="666666")
            return start_row + 2
        
        # Headers
        headers = ['Client Name', 'Refresh Status'] + [
            FIELD_DISPLAY_NAMES.get(field, field) for field in all_fields
        ] + ['Error Details']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write data
        current_row = start_row + 1
        
        for result in results:
            report_data = result[report_key]
            extracted_data = report_data.get('data', {})
            
            # Client name
            ws.cell(row=current_row, column=1, value=result['client']).border = border
            
            # Status
            status_cell = ws.cell(
                row=current_row, 
                column=2, 
                value='✓ Success' if report_data['status']['success'] else '✗ Failed'
            )
            status_cell.border = border
            if report_data['status']['success']:
                status_cell.font = Font(color="107C10", bold=True)
            else:
                status_cell.font = Font(color="D83B01", bold=True)
            
            # Data fields
            for col, field in enumerate(all_fields, 3):
                value = extracted_data.get(field)
                
                if value is None:
                    display_value = '-'
                elif isinstance(value, (int, float)):
                    display_value = value
                else:
                    display_value = str(value)
                
                cell = ws.cell(row=current_row, column=col, value=display_value)
                cell.border = border
                
                # Format numbers
                if isinstance(value, (int, float)) and value is not None:
                    cell.number_format = '#,##0.00'
            
            # Error details
            error_msg = report_data['status'].get('error', '-')
            error_cell = ws.cell(
                row=current_row, 
                column=len(headers),
                value=error_msg if error_msg else '-'
            )
            error_cell.border = border
            if error_msg and error_msg != '-':
                error_cell.font = Font(color="D83B01", size=9)
            
            current_row += 1
        
        # Add summary row
        current_row += 1
        ws.cell(row=current_row, column=1, value="Summary:").font = subsection_font
        
        success_count = sum(1 for r in results if r[report_key]['status']['success'])
        total_count = len(results)
        
        ws.cell(
            row=current_row, 
            column=2, 
            value=f"{success_count}/{total_count} successful"
        ).font = Font(italic=True)
        
        return current_row
    
    def _create_details_sheet(self, wb, results):
        """Create a detailed sheet showing all extraction info"""
        ws = wb.create_sheet("Extraction Details")
        
        # Title
        ws['A1'] = "Detailed Extraction Information"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        for result in results:
            # Client header
            ws[f'A{row}'] = f"Client: {result['client']}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # ITC Report
            ws[f'A{row}'] = "ITC Report:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            itc_data = result['itc']
            ws[f'B{row}'] = f"Status: {'Success' if itc_data['status']['success'] else 'Failed'}"
            row += 1

            if itc_data['status'].get('validation_message'):
                ws[f'B{row}'] = f"Validation: {itc_data['status']['validation_message']}"
                row += 1
            
            if itc_data.get('data'):
                ws[f'B{row}'] = "Extracted Values:"
                row += 1
                for field, value in itc_data['data'].items():
                    ws[f'C{row}'] = f"{field}: {value}"
                    row += 1
            
            row += 1
            
            # Sales Report
            ws[f'A{row}'] = "Sales Report:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            sales_data = result['sales']
            ws[f'B{row}'] = f"Status: {'Success' if sales_data['status']['success'] else 'Failed'}"
            row += 1

            if sales_data['status'].get('validation_message'):
                ws[f'B{row}'] = f"Validation: {sales_data['status']['validation_message']}"
                row += 1
            
            if sales_data.get('data'):
                ws[f'B{row}'] = "Extracted Values:"
                row += 1
                for field, value in sales_data['data'].items():
                    ws[f'C{row}'] = f"{field}: {value}"
                    row += 1
            
            row += 2
    
    def _autofit_columns(self, ws):
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                except (TypeError, ValueError):
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > 0:
                ws.column_dimensions[column_letter].width = adjusted_width