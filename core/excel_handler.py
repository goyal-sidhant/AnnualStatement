"""
Excel Handler for GST File Organizer v3.0
WINDOWS VERSION - Uses Excel COM to preserve Power Query and all Excel features
"""

import os
import time
import shutil
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import warnings
import platform

# Suppress warnings
warnings.filterwarnings('ignore')

# Check if we're on Windows and can use win32com
USE_WIN32COM = False
if platform.system() == 'Windows':
    try:
        import win32com.client
        import pythoncom
        USE_WIN32COM = True
        logging.info("Using win32com for Excel operations (Power Query support)")
    except ImportError:
        logging.warning("win32com not available. Install with: pip install pywin32")

# Fallback to openpyxl if not on Windows
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.constants import EXCEL_TEMPLATE_MAPPING, EXCEL_SAFETY
from utils.helpers import (
    sanitize_filename, get_safe_timestamp, ensure_path_exists,
    validate_excel_file, clean_windows_path, extract_filename_without_extension,
    get_short_path, get_state_code
)

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Excel handler with Power Query support on Windows
    """
    
    def __init__(self):
        """Initialize Excel handler"""
        self.temp_dir = Path(tempfile.gettempdir()) / "gst_excel_temp"
        ensure_path_exists(self.temp_dir)
        self.safety_config = EXCEL_SAFETY
        self.excel_app = None
        
    def __del__(self):
        """Cleanup Excel COM object"""
        if USE_WIN32COM and self.excel_app:
            try:
                self.excel_app.Quit()
            except Exception:
                pass
                
    def create_report_from_template(self, template_path: Union[str, Path],
                                    output_path: Union[str, Path],
                                    data_mappings: Dict[str, str],
                                    report_type: str) -> bool:
        """
        Create report from template preserving Power Query
        """
        logger.debug(f"Creating {report_type} report: template={template_path}, output={output_path}")

        template_path = Path(template_path)
        output_path = Path(output_path)

        # Ensure paths exist
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        ensure_path_exists(output_path.parent)
        
        if USE_WIN32COM:
            # Use Excel COM for Windows (preserves Power Query)
            return self._create_report_win32com(
                template_path, output_path, data_mappings, report_type
            )
        else:
            # Fallback to openpyxl (no Power Query support)
            logger.warning("Not on Windows - Power Query may not be preserved!")
            return self._create_report_openpyxl(
                template_path, output_path, data_mappings, report_type
            )
    
    def _create_report_win32com(self, template_path: Path, output_path: Path,
                               data_mappings: Dict[str, str], report_type: str) -> bool:
        """
        Create report using Excel COM (preserves all Excel features including Power Query)
        """
        logger.debug(f"Creating {report_type} report via win32com")

        excel = None
        wb = None
        
        try:
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Get Excel application
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                logger.info("Using existing Excel instance")
            except Exception:
                excel = win32com.client.Dispatch("Excel.Application")
                logger.info("Created new Excel instance")
            
            # Configure Excel
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            
            # First, copy template to output location
            logger.info(f"Copying template to output location...")
            shutil.copy2(str(template_path), str(output_path))
            time.sleep(0.5)  # Give Windows time to complete the copy
            
            # Open the copied file
            logger.info(f"Opening workbook: {output_path}")
            wb = self._safe_open_workbook(excel, output_path)
            
            # Get mapping configuration
            mapping_config = EXCEL_TEMPLATE_MAPPING.get(report_type, {})
            sheet_name = mapping_config.get('sheet', 'Links')
            
            # Find the Links sheet
            ws = None
            for sheet in wb.Sheets:
                if sheet.Name.lower() == sheet_name.lower():
                    ws = sheet
                    break
            
            if not ws:
                # Try to find any sheet with 'link' in name
                for sheet in wb.Sheets:
                    if 'link' in sheet.Name.lower():
                        ws = sheet
                        logger.warning(f"Using sheet '{sheet.Name}' instead of '{sheet_name}'")
                        break
            
            if not ws:
                logger.error(f"Links sheet not found! Creating one...")
                ws = wb.Sheets.Add()
                ws.Name = sheet_name
            
            # Update cells with data
            logger.info(f"Updating cells in {report_type} report:")
            cells_updated = 0
            
            for cell_ref, data_key in mapping_config.get('cells', {}).items():
                try:
                    value = data_mappings.get(data_key, '')
                    if value:
                        ws.Range(cell_ref).Value = str(value)
                        cells_updated += 1
                        logger.info(f"  {cell_ref} = {value[:50]}...")
                except Exception as e:
                    logger.error(f"  Failed to update {cell_ref}: {e}")
            
            logger.info(f"Updated {cells_updated} cells")
            
            # Save the workbook
            logger.info(f"Saving workbook...")
            wb.Save()
            
            # Close workbook
            wb.Close(SaveChanges=True)
            wb = None
            
            # Small delay
            time.sleep(0.5)
            
            # Verify the file
            if output_path.exists():
                file_size = output_path.stat().st_size
                logger.info(f"Successfully Created {report_type} report: {output_path.name} ({file_size} bytes)")
                return True
            else:
                logger.error("Output file not found!")
                return False
            
        except Exception as e:
            logger.error(f"Error creating report with win32com: {e}", exc_info=True)
            return False
            
        finally:
            # Cleanup
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass

            if excel:
                excel.ScreenUpdating = True
                excel.DisplayAlerts = True
                # Don't quit if Excel was already open
                if self.excel_app is None:
                    try:
                        excel.Quit()
                    except Exception:
                        pass

            # Uninitialize COM
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    
    def _create_report_openpyxl(self, template_path: Path, output_path: Path,
                               data_mappings: Dict[str, str], report_type: str) -> bool:
        """
        Fallback method using openpyxl (may lose Power Query)
        """
        logger.warning("Using openpyxl - Power Query content may be lost!")
        
        try:
            # Copy template first
            shutil.copy2(str(template_path), str(output_path))
            
            # Load and update
            wb = load_workbook(str(output_path), data_only=False, keep_links=True)
            
            # Get mapping
            mapping_config = EXCEL_TEMPLATE_MAPPING.get(report_type, {})
            sheet_name = mapping_config.get('sheet', 'Links')
            
            # Find sheet
            ws = None
            for sname in wb.sheetnames:
                if sname.lower() == sheet_name.lower():
                    ws = wb[sname]
                    break
            
            if ws:
                # Update cells
                for cell_ref, data_key in mapping_config.get('cells', {}).items():
                    try:
                        value = data_mappings.get(data_key, '')
                        if value:
                            ws[cell_ref] = str(value)
                    except (KeyError, ValueError) as e:
                        logger.warning(f"Could not update cell {cell_ref}: {e}")
            
            # Save
            wb.save(str(output_path))
            wb.close()
            
            return output_path.exists()
            
        except Exception as e:
            logger.error(f"Fallback method error: {e}")
            return False
    
    def _safe_open_workbook(self, excel, file_path):
        """Safely open workbook with short path fallback"""
        file_path = Path(file_path).absolute()
        
        # Check if file exists
        if not file_path.exists():
            logger.error(f"File does not exist: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Check if already open in Excel
        for wb in excel.Workbooks:
            if Path(wb.FullName).absolute() == file_path:
                logger.info(f"File already open, using existing workbook")
                return wb
        
        file_str = str(file_path)
        path_length = len(file_str)
        logger.info(f"Path length: {path_length} characters")
        
        # First attempt - standard open
        try:
            logger.info(f"Opening workbook: {file_path.name}")
            return excel.Workbooks.Open(file_str)
        except Exception as e:
            error_code = getattr(e, 'args', [None])[0] if hasattr(e, 'args') else None
            logger.error(f"First attempt failed - Error code: {error_code}")
            
            # Check if it's a path-related error
            if ("-2147352567" in str(e) or "path" in str(e).lower() or 
                path_length > 218):
                
                # Try short path approach
                try:
                    logger.info("Trying 8.3 short path...")
                    short_path = get_short_path(file_path)
                    if short_path != str(file_path):
                        logger.info(f"Using short path: ...{short_path[-50:]}")
                        return excel.Workbooks.Open(short_path)
                except Exception as short_e:
                    logger.error(f"Short path also failed: {short_e}")
            
            # Re-raise original error if short path doesn't help
            raise
    
    def create_summary_report(self, output_path: Union[str, Path],
                            report_data: Dict[str, Any]) -> bool:
        """
        Create summary report (doesn't need Power Query)
        """
        try:
            output_path = Path(output_path)
            ensure_path_exists(output_path.parent)
            
            # Always use openpyxl for summary (no Power Query needed)
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define styles
            styles = self._create_styles()
            
            # Create sheets
            self._create_summary_sheet(wb, report_data, styles)
            self._create_client_status_sheet(wb, report_data, styles)
            self._create_file_mapping_sheet(wb, report_data, styles)
            self._create_errors_sheet(wb, report_data, styles)
            self._create_variations_sheet(wb, report_data, styles)
            
            # Auto-fit columns
            for ws in wb.worksheets:
                self._autofit_columns(ws)
            
            # Save
            wb.save(str(output_path))
            wb.close()
            
            return output_path.exists()
            
        except Exception as e:
            logger.error(f"Error creating summary report: {e}", exc_info=True)
            return False
    
    def _create_styles(self) -> Dict[str, Any]:
        """Create Excel styles"""
        return {
            'header': {
                'font': Font(bold=True, size=12, color="FFFFFF"),
                'fill': PatternFill(start_color="366092", end_color="366092", 
                                  fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'title': {
                'font': Font(bold=True, size=16),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict, styles: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "GST FILE PROCESSING SUMMARY"
        ws['A1'].font = styles['title']['font']
        ws.merge_cells('A1:B1')
        
        # Data rows
        summary_data = [
            ("Processing Date", data.get('timestamp', '')),
            ("Total Clients", data.get('total_clients', 0)),
            ("Successful", data.get('successful_clients', 0)),
            ("Failed", data.get('failed_clients', 0)),
            ("Total Files", data.get('total_files', 0)),
            ("Reports Generated", data.get('reports_generated', 0)),
            ("ITC Reports Created", data.get('itc_reports_created', 0)),  # NEW
            ("Sales Reports Created", data.get('sales_reports_created', 0)),  # NEW
            ("Report Creation Errors", data.get('report_errors', 0)),  # NEW
            ("Processing Mode", data.get('processing_mode', '')),
            ("Target Folder", data.get('target_folder', '')),
            ("Include Client Name in Folders", "Yes" if data.get('include_client_name', False) else "No")
        ]
        
        row = 3
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_client_status_sheet(self, wb: Workbook, data: Dict, styles: Dict):
        """Create client status sheet"""
        ws = wb.create_sheet("Client Status")
        
        # Headers
        headers = ["Client", "State", "Status", "Files", "Missing", "Complete %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header']['font']
            cell.fill = styles['header']['fill']
            cell.alignment = styles['header']['alignment']
            cell.border = styles['border']
        
        # Data
        row = 2
        for client_data in data.get('clients', []):
            ws.cell(row=row, column=1, value=client_data.get('client', ''))
            ws.cell(row=row, column=2, value=client_data.get('state', ''))
            ws.cell(row=row, column=3, value=client_data.get('status', ''))
            ws.cell(row=row, column=4, value=client_data.get('file_count', 0))
            ws.cell(row=row, column=5, value=len(client_data.get('missing_files', [])))
            ws.cell(row=row, column=6, value=f"{client_data.get('completeness', 0):.1f}%")
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = styles['border']
            
            row += 1
    
    def _create_file_mapping_sheet(self, wb: Workbook, data: Dict, styles: Dict):
        """Create file mapping sheet"""
        ws = wb.create_sheet("File Mapping")
        
        # Headers
        headers = ["Original File", "Client", "Type", "Destination", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header']['font']
            cell.fill = styles['header']['fill']
            cell.alignment = styles['header']['alignment']
            cell.border = styles['border']
        
        # Data
        row = 2
        for file_map in data.get('file_mappings', []):
            ws.cell(row=row, column=1, value=file_map.get('filename', ''))
            ws.cell(row=row, column=2, value=file_map.get('client', ''))
            ws.cell(row=row, column=3, value=file_map.get('type', ''))
            ws.cell(row=row, column=4, value=file_map.get('destination', ''))
            ws.cell(row=row, column=5, value=file_map.get('status', ''))
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = styles['border']
            
            row += 1
    
    def _create_errors_sheet(self, wb: Workbook, data: Dict, styles: Dict):
        """Create errors sheet"""
        ws = wb.create_sheet("Errors")
        
        # Headers
        headers = ["Time", "Client", "Type", "Error", "Action"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header']['font']
            cell.fill = styles['header']['fill']
            cell.alignment = styles['header']['alignment']
            cell.border = styles['border']
        
        # Data
        row = 2
        errors = data.get('errors', [])
        if not errors:
            ws.cell(row=2, column=1, value="No errors encountered")
            ws.merge_cells('A2:E2')
        else:
            for error in errors:
                ws.cell(row=row, column=1, value=error.get('time', ''))
                ws.cell(row=row, column=2, value=error.get('client', ''))
                ws.cell(row=row, column=3, value=error.get('type', ''))
                ws.cell(row=row, column=4, value=error.get('message', ''))
                ws.cell(row=row, column=5, value=error.get('action', 'Review'))
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = styles['border']
                
                row += 1
    
    def _create_variations_sheet(self, wb: Workbook, data: Dict, styles: Dict):
        """Create variations sheet"""
        ws = wb.create_sheet("Variations")
        
        # Headers
        headers = ["Filename", "Issue", "Expected Pattern", "Size", "Action"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header']['font']
            cell.fill = styles['header']['fill']
            cell.alignment = styles['header']['alignment']
            cell.border = styles['border']
        
        # Data
        row = 2
        variations = data.get('variations', [])
        if not variations:
            ws.cell(row=2, column=1, value="No file variations found")
            ws.merge_cells('A2:E2')
        else:
            for var in variations:
                ws.cell(row=row, column=1, value=var.get('filename', ''))
                ws.cell(row=row, column=2, value=var.get('issue', ''))
                ws.cell(row=row, column=3, value=var.get('pattern_expected', ''))
                ws.cell(row=row, column=4, value=var.get('size', ''))
                ws.cell(row=row, column=5, value="Rename file")
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = styles['border']
                
                row += 1
    
    def _autofit_columns(self, ws):
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def prepare_template_data(self, client_info: Dict[str, Any],
                            folders: Dict[str, Path],
                            report_type: str) -> Dict[str, str]:
        """
        Prepare data mappings for template
        """
        logger.debug(f"Preparing {report_type} template data for {client_info.get('client', 'Unknown')}")

        data = {}

        state_code = get_state_code(client_info['state'])

        try:
            if report_type == 'ITC':
                # GSTR-3B folder
                data['gstr3b_folder'] = clean_windows_path(folders.get('gstr3b', ''))
                
                # Annual report (in version folder)
                annual_files = client_info['files'].get('Annual Report', [])
                if annual_files:
                    data['annual_folder'] = clean_windows_path(folders.get('version', ''))
                    data['annual_filename'] = extract_filename_without_extension(
                        annual_files[0]['name']
                    )
                else:
                    data['annual_folder'] = ''
                    data['annual_filename'] = ''
                
                # GSTR-2B Reco
                gstr2b_files = client_info['files'].get('GSTR-2B Reco', [])
                if gstr2b_files:
                    data['gstr2b_folder'] = clean_windows_path(folders.get('itc', ''))
                    data['gstr2b_filename'] = extract_filename_without_extension(
                        gstr2b_files[0]['name']
                    )
                else:
                    data['gstr2b_folder'] = ''
                    data['gstr2b_filename'] = ''
                
                # IMS Reco
                ims_files = client_info['files'].get('IMS Reco', [])
                if ims_files:
                    data['ims_folder'] = clean_windows_path(folders.get('itc', ''))
                    data['ims_filename'] = extract_filename_without_extension(
                        ims_files[0]['name']
                    )
                else:
                    data['ims_folder'] = ''
                    data['ims_filename'] = ''
                    
            elif report_type == 'Sales':
                # Sales folder
                sales_files = client_info['files'].get('Sales', [])
                if sales_files:
                    data['sales_folder'] = clean_windows_path(folders.get('sales', ''))
                    data['sales_filename'] = extract_filename_without_extension(
                        sales_files[0]['name']
                    )
                else:
                    data['sales_folder'] = ''
                    data['sales_filename'] = ''

                # Annual report (in version folder)
                annual_files = client_info['files'].get('Annual Report', [])
                if annual_files:
                    data['annual_folder'] = clean_windows_path(folders.get('version', ''))
                    data['annual_filename'] = extract_filename_without_extension(
                        annual_files[0]['name']
                    )
                else:
                    data['annual_folder'] = ''
                    data['annual_filename'] = ''

                # Sales Reco
                sales_reco_files = client_info['files'].get('Sales Reco', [])
                if sales_reco_files:
                    data['sales_reco_folder'] = clean_windows_path(folders.get('sales', ''))
                    data['sales_reco_filename'] = extract_filename_without_extension(
                        sales_reco_files[0]['name']
                    )
                else:
                    data['sales_reco_folder'] = ''
                    data['sales_reco_filename'] = ''
            
        except Exception as e:
            logger.error(f"Error preparing template data: {e}")
        
        return data